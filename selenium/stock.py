import shutil
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import date
import urllib.request
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# Copy the format file
stock_dir = "C:/Users/james/Desktop/Self/Stock/"
stock_format = "종목선정_분석표_양식.xlsx"

# new_stock = input("Enter stock name: ")
new_stock = "LG우"
new_file_name = "종목선정_분석표_" +  new_stock +".xlsx"
# shutil.copy(stock_dir + stock_format, stock_dir + new_file_name)

# Load the workbook and worksheet
load_wb = load_workbook(filename = stock_dir + new_file_name)
load_ws = load_wb.worksheets[0]

a2 = 'A2'
today = date.today().strftime("%d/%m/%Y")
load_ws[a2] = "(종목명 : " + new_stock + "  일시 : " + today + ")"
load_ws[a2].alignment = Alignment(horizontal='right')

# Crawl financial data
options = Options()
options.add_argument("--incognito")

link = "https://finance.naver.com/"
driver = webdriver.Chrome(options=options)
driver.get(link)

searchBar = driver.find_element_by_xpath('//*[@id="stock_items"]')
searchBar.send_keys(new_stock)
searchButton = driver.find_element_by_xpath('//*[@id="header"]/div[1]/div/div[2]/form/fieldset/div/button')
searchButton.click()
time.sleep(1)

# Net Income
net_income_year = []
net_income_quarter = []
for i in range(1, 5):
    net_income_year.append(int(driver.find_element_by_xpath('//*[@id="content"]/div[4]/div[1]/table/tbody/tr[3]/td[' + str(i) + ']').text.replace(',', '') ))
    net_income_quarter.append(int(driver.find_element_by_xpath('//*[@id="content"]/div[4]/div[1]/table/tbody/tr[3]/td[' + str(i + 5) + ']').text.replace(',', '') ))
    load_ws.cell(row = 5, column = i * 2).value = net_income_year[i - 1]
    load_ws.cell(row = 7, column = i * 2).value = net_income_quarter[i - 1]
    # load_ws.cell(row = 5, column = i * 2).style = '#,##0'
    # load_ws.cell(row = 7, column = i * 2).style = '#,##0'

net_income_quarter.append(int(driver.find_element_by_xpath('//*[@id="content"]/div[4]/div[1]/table/tbody/tr[3]/td[10]').text))
load_ws.cell(row = 7, column = 10).value = net_income_quarter[4]

# Aggregate value
num_of_stocks = int(driver.find_element_by_xpath('//*[@id="tab_con1"]/div[1]/table/tbody/tr[3]/td/em').text.replace(',', '') )
price_of_stock = int(driver.find_element_by_xpath('//*[@id="chart_area"]/div[1]/div/p[1]').text.replace('\n', '').replace(',', ''))
aggregate_value = int(driver.find_element_by_xpath('//*[@id="_market_sum"]').text.replace(',', '') )

load_ws.cell(row = 13, column = 2).value = num_of_stocks
load_ws.cell(row = 13, column = 5).value = price_of_stock
load_ws.cell(row = 13, column = 8).value = aggregate_value

# Save the file
load_wb.save(filename = stock_dir + new_file_name)

driver.quit()

#TODO
# Format numbers, fill in rest of the values